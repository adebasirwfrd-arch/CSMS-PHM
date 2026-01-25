
import csv
import pdfplumber
import openpyxl
from docx import Document
from pptx import Presentation
import io
import os
import re
from typing import Dict, List, Any, Optional

class ReportEngine:
    """
    Core engine to parse input data (Excel, PDF) and fill templates (Word, Excel, PPT).
    Replaced pandas with native libraries to reduce serverless package size.
    """

    def __init__(self):
        pass

    def parse_excel_source(self, file_content: bytes) -> List[Dict[str, Any]]:
        """
        Parse an Excel source file using openpyxl instead of pandas
        """
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            ws = wb.active
            
            data = []
            headers = []
            
            # Get headers from first row
            for cell in ws[1]:
                headers.append(str(cell.value).strip() if cell.value else f"col_{cell.col_idx}")
                
            # Parse rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                record = {}
                has_data = False
                for i, value in enumerate(row):
                    if i < len(headers):
                        record[headers[i]] = value
                        if value: has_data = True
                
                if has_data:
                    data.append(record)
                    
            return data
        except Exception as e:
            print(f"[ReportEngine] Error parsing Excel: {e}")
            return []

    def parse_pdf_source(self, file_content: bytes) -> Dict[str, Any]:
        """
        Extract data from PDF using pdfplumber.
        Returns a dictionary with 'employee_name' and 'records' list.
        Handles multiple table formats:
        1. Competency Role History table (Name, Start Date, End Date, Compliance)
        2. Training table (Name, External ID, Type, Pass Rate, Completion Date)
        """
        data = {
            'employee_name': None,
            'records': []
        }
        
        try:
            with pdfplumber.open(io.BytesIO(file_content)) as pdf:
                # 1. Try to extract Employee Name from first page text
                first_page_text = pdf.pages[0].extract_text()
                if first_page_text:
                    lines = [l.strip() for l in first_page_text.split('\n') if l.strip()]
                    if lines:
                        data['employee_name'] = lines[0] 
                        
                # 2. Extract Tables from ALL pages
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        if not table or len(table) < 2: continue
                        
                        # Get header row (clean up None values)
                        header_row = [str(h).lower().replace('\n', ' ').strip() if h else '' for h in table[0]]
                        
                        # Check for Competency Role History table (has 'compliance')
                        if 'name' in header_row and 'compliance' in header_row:
                            try:
                                name_idx = header_row.index('name')
                                start_idx = next((i for i, h in enumerate(header_row) if 'start' in h), None)
                                end_idx = next((i for i, h in enumerate(header_row) if 'end' in h), None)
                                
                                for row in table[1:]:
                                    if len(row) <= name_idx or not row[name_idx]: continue
                                    record = {
                                        'Training Name': str(row[name_idx]).strip(),
                                        'Start Date': row[start_idx] if start_idx is not None and len(row) > start_idx else None,
                                        'End Date': row[end_idx] if end_idx is not None and len(row) > end_idx else None
                                    }
                                    if record['Training Name']:
                                        data['records'].append(record)
                            except (IndexError, ValueError):
                                continue
                        
                        # Check for Training table (has 'completion date')
                        elif 'name' in header_row and any('completion' in h for h in header_row):
                            try:
                                name_idx = header_row.index('name')
                                completion_idx = next((i for i, h in enumerate(header_row) if 'completion' in h), None)
                                
                                for row in table[1:]:
                                    if len(row) <= name_idx or not row[name_idx]: continue
                                    training_name = str(row[name_idx]).strip()
                                    completion_date = row[completion_idx] if completion_idx is not None and len(row) > completion_idx else None
                                    
                                    record = {
                                        'Training Name': training_name,
                                        'Start Date': completion_date,  # Use completion date as training date
                                        'End Date': None  # No expiry in this table
                                    }
                                    if record['Training Name']:
                                        data['records'].append(record)
                            except (IndexError, ValueError):
                                continue
                
                # Also try to extract from page text if tables didn't work well
                # Look for "Training" section in page 2
                for page in pdf.pages:
                    text = page.extract_text()
                    if text and 'Training' in text:
                        lines = text.split('\n')
                        for line in lines:
                            # Look for lines matching "RightStart - ..." pattern
                            if 'RightStart' in line or 'WFRD CORE' in line or 'GEOZONE' in line:
                                # Try to parse: Name ID Type Pass% Date
                                parts = line.split()
                                # Find date pattern (MM/DD/YYYY)
                                date_match = None
                                for part in parts:
                                    if '/' in part and len(part) >= 8:
                                        date_match = part
                                        break
                                if date_match:
                                    # Get training name (everything before the ID number)
                                    name_parts = []
                                    for part in parts:
                                        if part.isdigit() and len(part) == 4:
                                            break
                                        name_parts.append(part)
                                    training_name = ' '.join(name_parts)
                                    
                                    # Check if we already have this record
                                    existing = [r for r in data['records'] if r['Training Name'].lower() == training_name.lower()]
                                    if not existing and training_name:
                                        data['records'].append({
                                            'Training Name': training_name,
                                            'Start Date': date_match,
                                            'End Date': None
                                        })
                
                print(f"[ReportEngine] Parsed {len(data['records'])} training records for {data['employee_name']}")
                
            return data
        except Exception as e:
            print(f"[ReportEngine] PDF Parse Error: {e}")
            import traceback
            traceback.print_exc()
            return data

    def fill_matrix_template(self, template_content: bytes, source_data_list: List[Dict[str, Any]]) -> bytes:
        """
        Specialized filler for Training Matrix - Handles MULTIPLE Employees.
        """
        try:
            wb = openpyxl.load_workbook(io.BytesIO(template_content))
            ws = wb.active
            
            # 1. Map Headers (Do this once)
            training_map = {}
            header_row = 2
            sub_header_row = 3
            current_training = None
            
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=col).value
                if val:
                    current_training = str(val).strip()
                    if current_training not in training_map:
                        training_map[current_training] = {}
                
                sub_val = ws.cell(row=sub_header_row, column=col).value
                if current_training and sub_val:
                    sub_val_str = str(sub_val).strip().lower()
                    if 'training date' == sub_val_str:
                         training_map[current_training]['date_col'] = col
                    elif 'expiry' in sub_val_str:
                         training_map[current_training]['expiry_col'] = col

            print(f"[ReportEngine] Processing {len(source_data_list)} employees for Matrix...")

            # 2. Iterate through EACH employee source
            for source_data in source_data_list:
                employee_name = source_data.get('employee_name')
                records = source_data.get('records', [])
                
                if not employee_name:
                    continue
                    
                print(f"  -> Employee: {employee_name}")
                
                # Find Employee Row
                emp_row_idx = None
                name_col_idx = 3 
                for row in ws.iter_rows(min_row=4, max_col=5):
                    cell_val = row[name_col_idx-1].value
                    if cell_val and (str(employee_name).lower() in str(cell_val).lower() or str(cell_val).lower() in str(employee_name).lower()):
                        emp_row_idx = row[0].row
                        break
                
                if not emp_row_idx:
                    print(f"     [!] Row not found in Excel.")
                    continue

                # Fill Data
                for record in records:
                    t_name = record.get('Training Name')
                    t_start = record.get('Start Date')
                    t_end = record.get('End Date')
                    
                    if not t_name: continue
                    
                    # Normalize training name from PDF
                    t_name_clean = t_name.lower().strip()

                    matched = False
                    for header, cols in training_map.items():
                        header_clean = header.lower().strip()
                        
                        # Exact match (case-insensitive)
                        if t_name_clean == header_clean:
                            matched = True
                        # Also match if PDF training contains the template header exactly
                        elif header_clean in t_name_clean:
                            matched = True
                        
                        if matched:
                            if 'date_col' in cols and t_start:
                                ws.cell(row=emp_row_idx, column=cols['date_col']).value = t_start
                            if 'expiry_col' in cols and t_end:
                                ws.cell(row=emp_row_idx, column=cols['expiry_col']).value = t_end
                            break
            
            # 3. Fill empty cells with "N/A"
            for row_idx in range(4, ws.max_row + 1):
                emp_name = ws.cell(row=row_idx, column=3).value
                if not emp_name: continue
                
                for header, cols in training_map.items():
                    if 'date_col' in cols:
                        cell = ws.cell(row=row_idx, column=cols['date_col'])
                        if cell.value is None or str(cell.value).strip() == '':
                            cell.value = 'N/A'
                    if 'expiry_col' in cols:
                        cell = ws.cell(row=row_idx, column=cols['expiry_col'])
                        if cell.value is None or str(cell.value).strip() == '':
                            cell.value = 'N/A'

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()
            
        except Exception as e:
            print(f"[ReportEngine] Matrix Fill Error: {e}")
            return template_content

    def fill_csv_template(self, data_records: List[Dict[str, Any]]) -> bytes:
        """
        Generate a CSV from data records using csv module (removed pandas)
        """
        try:
            if not data_records:
                return b"No data found"
                
            output = io.StringIO()
            if data_records:
                headers = data_records[0].keys()
                writer = csv.DictWriter(output, fieldnames=headers)
                writer.writeheader()
                writer.writerows(data_records)
                
            return output.getvalue().encode('utf-8')
        except Exception as e:
            print(f"[ReportEngine] Error generating CSV: {e}")
            return b"Error generating CSV"


    def fill_excel_template(self, template_content: bytes, data_records: List[Dict[str, Any]]) -> bytes:
        """
        Fill a flexible Excel template by appending rows
        """
        try:
            wb = openpyxl.load_workbook(io.BytesIO(template_content))
            ws = wb.active
            
            # Append data based on headers in first row or just list dict keys
            headers = []
            if ws.max_row >= 1:
                headers = [str(cell.value) for cell in ws[1] if cell.value]
            
            # If no headers in template, use keys from first record
            if not headers and data_records:
                headers = list(data_records[0].keys())
                ws.append(headers)
            
            for record in data_records:
                row = []
                for h in headers:
                    row.append(record.get(h, ''))
                ws.append(row)
                
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()
        except Exception as e:
            print(f"[ReportEngine] Excel Fill Error: {e}")
            return template_content

    def process_request(self, template_content: bytes, template_filename: str, source_contents: List[bytes], source_filenames: List[str]) -> bytes:
        """
        Main entry point. Handles multiple sources.
        """
        # 1. Parse All Sources
        all_source_data = []
        
        for content, filename in zip(source_contents, source_filenames):
            if filename.lower().endswith('.xlsx') or filename.lower().endswith('.xls'):
                data = self.parse_excel_source(content)
                if isinstance(data, list):
                    all_source_data.append({'records': data, 'filename': filename})
                else:
                    all_source_data.append(data)
                    
            elif filename.lower().endswith('.pdf'):
                data = self.parse_pdf_source(content)
                all_source_data.append(data)
            else:
                print(f"[ReportEngine] Skip unsupported: {filename}")

        if not all_source_data:
            print("[ReportEngine] No data extracted from sources")
            return None

        # 2. Fill Template
        generated_file = None
        
        # Matrix Check
        is_matrix = False
        if template_filename.lower().endswith('.xlsx'):
            if any(d.get('employee_name') for d in all_source_data if isinstance(d, dict)):
                is_matrix = True

        if template_filename.lower().endswith('.docx'):
            # Stub for docx (would need python-docx code here if used)
            print("[ReportEngine] DOCX not fully implemented in lightweight version")
            return None
            
        elif template_filename.lower().endswith('.xlsx'):
            if is_matrix:
                generated_file = self.fill_matrix_template(template_content, all_source_data)
            else:
                # Concatenate all records
                all_records = []
                for d in all_source_data:
                    recs = d.get('records', []) if isinstance(d, dict) else d
                    if isinstance(recs, list):
                         all_records.extend(recs)
                generated_file = self.fill_excel_template(template_content, all_records)
        
        elif template_filename.lower().endswith('.csv'):
             # Concatenate
            all_records = []
            for d in all_source_data:
                recs = d.get('records', []) if isinstance(d, dict) else d
                if isinstance(recs, list):
                    all_records.extend(recs)
            generated_file = self.fill_csv_template(all_records)
        
        else:
             print(f"[ReportEngine] Unsupported template: {template_filename}")
             return None
             
        return generated_file
