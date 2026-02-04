
import io
import os
from datetime import datetime
from typing import List, Dict
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    Workbook = None

class ExcelSyncService:
    def __init__(self, drive_service):
        self.drive_service = drive_service

    async def sync_to_drive(self, projects: List[Dict], tasks: List[Dict]):
        if not self.drive_service.enabled:
            print("[WARN] Drive not enabled, skipping Excel sync")
            return

        print("[INFO] Generating Excel report in memory...")
        file_data = self._generate_excel(projects, tasks)
        
        if not file_data:
            print("[ERROR] Failed to generate Excel data")
            return

        # Upload using the generic upload
        await self.drive_service.upload_file_to_drive(
            file_data=file_data,
            filename=f"CSMS_Track_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            project_name="CSMS_REPORTS"
        )
        
    def _generate_excel(self, projects: List[Dict], tasks: List[Dict]) -> bytes:
        if not Workbook:
            print("[ERROR] openpyxl not installed. Skipping Excel generation.")
            return None

        try:
            wb = Workbook()
            
            # Sheet 1: Projects
            ws_p = wb.active
            ws_p.title = "Projects"
            headers_p = ["ID", "Name", "Status", "Start Date", "End Date", "Description", "Created At"]
            ws_p.append(headers_p)
            
            for p in projects:
                ws_p.append([
                    p.get('id'), p.get('name'), p.get('status'), 
                    p.get('start_date'), p.get('end_date'), 
                    p.get('description'), p.get('created_at')
                ])

            # Sheet 2: Tasks
            ws_t = wb.create_sheet("Tasks")
            headers_t = ["Project ID", "Task Title", "Code", "Category", "Status", "Attachment Info"]
            ws_t.append(headers_t)
            
            for t in tasks:
                att_info = "Yes" if t.get('attachments') else "No"
                ws_t.append([
                    t.get('project_id'), t.get('title'), t.get('code'),
                    t.get('category'), t.get('status'), att_info
                ])

            # Save to memory instead of file
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
        except Exception as e:
            print(f"[ExcelSyncService] Error: {e}")
            return None
